from getopt import GetoptError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# For GUI Interface to get files path (Select csv file)
from getFilePath import App 

def getRollNumber(description):
    rollNo=(description.split(' '))[3]
    return rollNo[1:]

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Insert Header
ws.append(["Roll Number", "Txn Date","Description","Ref No./Check No.","Branch Code","Credit"])
for i in "ABCDEF":
    ws[i+'1'].font = Font(bold=True)
    
ws.title="Sheet 1"
ws.sheet_properties.tabColor = "1072BA"


""" Fill the values in this spreadsheet.  """

# Read the file with one record at a time

""""""
nap=App()
path=nap.getFilePath()

bankStatementWb = load_workbook(path)
bsWb = bankStatementWb.worksheets[0]


row_count = bsWb.max_row
print(row_count)

for i in range(2,row_count+1):
    description=bsWb[i][2].value
    valueDate=bsWb[i][1].value
    txnDate=bsWb[i][0].value
    refNo=bsWb[i][3].value
    branchCode=bsWb[i][4].value
    credit=bsWb[i][5].value

    currRollNumber=getRollNumber(description)
    ws.append([ currRollNumber, txnDate, description, refNo,branchCode, credit])


# Save the file
parentFolder=nap.getDirectoryPath()

wb.save(parentFolder+"\\"+"output.xlsx")
